"""Excel export for staff (spec 34), using openpyxl."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.utils.paths import EXPORTS_DIR

HEADER_FILL = PatternFill(start_color="0B3D66", end_color="0B3D66", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max([len(str(headers[col - 1]))] + [len(str(r.get(headers[col - 1], ""))) for r in rows])
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 40)
    ws.freeze_panes = "A2"


def export_rows(entity_name: str, headers: list[str], rows: list[dict]) -> Path:
    """Generic single-sheet export -- used for patients, appointments,
    consultations, prescriptions, lab reports, billing, payments, doctors."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = entity_name[:31]
    _write_sheet(ws, headers, rows)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = EXPORTS_DIR / f"{entity_name}_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


def export_analytics(patient_stats: dict, appointment_stats: dict, doctor_workload: list[dict],
                      financial_stats: dict) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Patients"
    ws1.append(["Metric", "Value"])
    for col in range(1, 3):
        ws1.cell(row=1, column=col).fill = HEADER_FILL
        ws1.cell(row=1, column=col).font = HEADER_FONT
    ws1.append(["Active Patients", patient_stats["total_active"]])
    ws1.append(["New (last 30 days)", patient_stats["new_last_30_days"]])
    ws1.append(["Returning Patients", patient_stats["returning"]])
    for gender, count in patient_stats["by_gender"].items():
        ws1.append([f"Gender: {gender}", count])

    ws2 = wb.create_sheet("Appointments")
    ws2.append(["Metric", "Value"])
    for col in range(1, 3):
        ws2.cell(row=1, column=col).fill = HEADER_FILL
        ws2.cell(row=1, column=col).font = HEADER_FONT
    ws2.append(["Total (period)", appointment_stats["total"]])
    for status, count in appointment_stats["by_status"].items():
        ws2.append([f"Status: {status}", count])

    ws3 = wb.create_sheet("Doctor Workload")
    _write_sheet(ws3, ["doctor_id", "full_name", "specialization", "appointment_count", "completed_count"],
                 doctor_workload)

    ws4 = wb.create_sheet("Financial")
    ws4.append(["Metric", "Value"])
    for col in range(1, 3):
        ws4.cell(row=1, column=col).fill = HEADER_FILL
        ws4.cell(row=1, column=col).font = HEADER_FONT
    ws4.append(["Revenue (period)", financial_stats["revenue_last_period"]])
    ws4.append(["Pending Amount", financial_stats["pending_amount"]])
    ws4.append(["Paid Invoices", financial_stats["paid_invoices"]])
    ws4.append(["Pending Invoices", financial_stats["pending_invoices"]])
    for category, total in financial_stats["revenue_by_category"].items():
        ws4.append([f"Category: {category}", total])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = EXPORTS_DIR / f"analytics_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path
